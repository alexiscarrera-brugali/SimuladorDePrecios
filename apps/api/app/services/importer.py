from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.domain.decimal_utils import decimal_to_str, parse_decimal, parse_percentage
from app.domain.types import Severity, SourceStatus

REQUIRED_COLUMNS = {
    "BD_LP": {"Sucursal", "Cod. Tabla", "Cod.Producto", "Precio Venta", "Vigencia", "Activo"},
    "SB1": {"Sucursal", "Codigo", "Descripcion", "Costo Estand", "Vigencia", "Desactivado?"},
    "Mapeo_Listas": {"Cod. Tabla", "Descripcion"},
    "Margen_teorico": {"Lista", "Código", "Descripción", "Margen"},
}


@dataclass
class ParsedIssue:
    issue_type: str
    severity: str
    sheet_name: str
    business_key: str
    explanation: str
    source_rows: list[int] = field(default_factory=list)
    values: list[str | None] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "issue_type": self.issue_type,
            "severity": self.severity,
            "sheet_name": self.sheet_name,
            "business_key": self.business_key,
            "explanation": self.explanation,
            "source_rows": self.source_rows,
            "values": self.values,
        }


@dataclass
class ParsedWorkbook:
    sha256: str
    raw_records: list[dict[str, Any]]
    price_lists: list[dict[str, Any]]
    prices: list[dict[str, Any]]
    costs: list[dict[str, Any]]
    margins: list[dict[str, Any]]
    issues: list[ParsedIssue]

    @property
    def summary(self) -> dict[str, int]:
        return {
            "price_rows": len(self.prices),
            "cost_rows": len(self.costs),
            "margin_rows": len(self.margins),
            "price_lists": len(self.price_lists),
            "warnings": sum(issue.severity == Severity.WARNING for issue in self.issues),
            "conflicts": sum(issue.severity == Severity.CONFLICT for issue in self.issues),
        }


def normalize_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return datetime.fromisoformat(value.strip()).date()
    raise ValueError(f"Invalid date: {value!r}")


def price_status(value: object) -> SourceStatus:
    normalized = str(value or "").strip().lower()
    if normalized in {"si", "sí", "s"}:
        return SourceStatus.ACTIVE
    if normalized in {"no", "n"}:
        return SourceStatus.INACTIVE
    return SourceStatus.UNKNOWN


def product_status(value: object) -> SourceStatus:
    normalized = str(value or "").strip().lower()
    if normalized in {"no", "n"}:
        return SourceStatus.ACTIVE
    if normalized in {"si", "sí", "s"}:
        return SourceStatus.INACTIVE
    return SourceStatus.UNKNOWN


def _records(sheet) -> tuple[list[str], list[tuple[Any, ...]]]:
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
    return headers, list(iterator)


def parse_workbook(path: Path) -> ParsedWorkbook:
    digest = sha256(path.read_bytes()).hexdigest()
    # En modo read_only openpyxl mantiene el archivo abierto; hay que cerrarlo
    # explícitamente o Windows bloquea el posterior unlink() del preview.
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, tuple[list[str], list[tuple[Any, ...]]]] = {}
    raw_records: list[dict[str, Any]] = []
    try:
        missing_sheets = set(REQUIRED_COLUMNS) - set(workbook.sheetnames)
        if missing_sheets:
            raise ValueError(f"Missing required sheets: {', '.join(sorted(missing_sheets))}")

        for sheet_name, required in REQUIRED_COLUMNS.items():
            headers, rows = _records(workbook[sheet_name])
            missing_columns = required - set(headers)
            if missing_columns:
                raise ValueError(
                    f"Sheet {sheet_name} is missing columns: {', '.join(sorted(missing_columns))}"
                )
            sheets[sheet_name] = (headers, rows)
            for source_row, row in enumerate(rows, start=2):
                payload = {
                    header: value.isoformat() if isinstance(value, (date, datetime)) else value
                    for header, value in zip(headers, row, strict=False)
                    if header
                }
                raw_records.append(
                    {"sheet_name": sheet_name, "source_row": source_row, "payload": payload}
                )
    finally:
        workbook.close()

    list_headers, list_rows = sheets["Mapeo_Listas"]
    list_index = {name: list_headers.index(name) for name in REQUIRED_COLUMNS["Mapeo_Listas"]}
    price_lists = [
        {
            "code": normalize_code(row[list_index["Cod. Tabla"]]),
            "description": str(row[list_index["Descripcion"]] or "").strip(),
        }
        for row in list_rows
        if row[list_index["Cod. Tabla"]] is not None
    ]

    issues: list[ParsedIssue] = []
    prices: list[dict[str, Any]] = []
    headers, rows = sheets["BD_LP"]
    index = {name: headers.index(name) for name in REQUIRED_COLUMNS["BD_LP"]}
    for source_row, row in enumerate(rows, start=2):
        try:
            record = {
                "branch_code": normalize_code(row[index["Sucursal"]]) or "1",
                "price_list_code": normalize_code(row[index["Cod. Tabla"]]),
                "product_code": normalize_code(row[index["Cod.Producto"]]),
                "value": parse_decimal(row[index["Precio Venta"]]),
                "valid_from": normalize_date(row[index["Vigencia"]]),
                "source_status": price_status(row[index["Activo"]]).value,
                "source_row": source_row,
            }
            prices.append(record)
        except ValueError as exc:
            issues.append(ParsedIssue("invalid_price_row", "conflict", "BD_LP", f"row:{source_row}", str(exc), [source_row]))

    costs: list[dict[str, Any]] = []
    headers, rows = sheets["SB1"]
    index = {name: headers.index(name) for name in REQUIRED_COLUMNS["SB1"]}
    for source_row, row in enumerate(rows, start=2):
        try:
            record = {
                "branch_code": normalize_code(row[index["Sucursal"]]) or "1",
                "product_code": normalize_code(row[index["Codigo"]]),
                "description": str(row[index["Descripcion"]] or "").strip() or None,
                "value": parse_decimal(row[index["Costo Estand"]]),
                "valid_from": normalize_date(row[index["Vigencia"]]),
                "source_status": product_status(row[index["Desactivado?"]]).value,
                "source_row": source_row,
            }
            costs.append(record)
        except ValueError as exc:
            issues.append(ParsedIssue("invalid_cost_row", "conflict", "SB1", f"row:{source_row}", str(exc), [source_row]))

    margins: list[dict[str, Any]] = []
    headers, rows = sheets["Margen_teorico"]
    index = {name: headers.index(name) for name in REQUIRED_COLUMNS["Margen_teorico"]}
    for source_row, row in enumerate(rows, start=2):
        product_code = normalize_code(row[index["Código"]])
        is_ambiguous = product_code.lower() == "varios"
        record = {
            "price_list_name": str(row[index["Lista"]] or "").strip(),
            "product_code": product_code,
            "percentage": parse_percentage(row[index["Margen"]]),
            "is_ambiguous": is_ambiguous,
            "source_row": source_row,
        }
        margins.append(record)
        if is_ambiguous:
            issues.append(
                ParsedIssue(
                    "objective_mapping_ambiguous",
                    "warning",
                    "Margen_teorico",
                    f"{record['price_list_name']}|varios",
                    "No se aplica hasta confirmar si 'varios' es un objetivo general.",
                    [source_row],
                    [decimal_to_str(record["percentage"])],
                )
            )

    issues.extend(_quality_issues(prices, "price"))
    issues.extend(_quality_issues(costs, "cost"))
    return ParsedWorkbook(digest, raw_records, price_lists, prices, costs, margins, issues)


def _quality_issues(records: list[dict[str, Any]], kind: str) -> list[ParsedIssue]:
    issues: list[ParsedIssue] = []
    grouped: dict[tuple, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        key = (
            record["branch_code"],
            record.get("price_list_code"),
            record["product_code"],
            record["valid_from"],
        )
        grouped[key].append(record)
        prefix = "price" if kind == "price" else "cost"
        if record["value"] is None:
            issues.append(ParsedIssue(f"missing_{prefix}", "warning", "BD_LP" if kind == "price" else "SB1", "|".join(map(str, key)), f"El {prefix} está vacío.", [record["source_row"]]))
        elif record["value"] == 0:
            issues.append(ParsedIssue(f"zero_{prefix}", "warning", "BD_LP" if kind == "price" else "SB1", "|".join(map(str, key)), f"El {prefix} es cero.", [record["source_row"]], ["0"]))
        elif record["value"] < 0:
            issues.append(ParsedIssue(f"negative_{prefix}", "warning", "BD_LP" if kind == "price" else "SB1", "|".join(map(str, key)), f"El {prefix} es negativo.", [record["source_row"]], [decimal_to_str(record["value"])]))
        if record["source_status"] == "inactive":
            issues.append(ParsedIssue("inactive_source", "warning", "BD_LP" if kind == "price" else "SB1", "|".join(map(str, key)), "El registro está marcado como inactivo.", [record["source_row"]]))
        elif record["source_status"] == "unknown":
            issues.append(ParsedIssue("unknown_source_status", "warning", "BD_LP" if kind == "price" else "SB1", "|".join(map(str, key)), "El estado está vacío o no reconocido; no se interpreta como inactivo.", [record["source_row"]]))

    for key, matches in grouped.items():
        if len(matches) < 2:
            continue
        distinct = {item["value"] for item in matches}
        conflicting = len(distinct) > 1
        issues.append(
            ParsedIssue(
                "conflicting_duplicate" if conflicting else "identical_duplicate",
                "conflict" if conflicting else "warning",
                "BD_LP" if kind == "price" else "SB1",
                "|".join(map(str, key)),
                "Hay valores diferentes para la misma clave y fecha; la fila queda bloqueada."
                if conflicting
                else "La misma clave, fecha y valor aparece más de una vez; se consolida para calcular.",
                [item["source_row"] for item in matches],
                [decimal_to_str(item["value"]) for item in matches],
            )
        )
    return issues

