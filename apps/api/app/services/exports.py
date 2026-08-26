from datetime import UTC, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.base import AuditEvent, QualityIssue, User
from app.domain.decimal_utils import decimal_to_str
from app.domain.simulation import simulate
from app.domain.types import Driver
from app.schemas.api import ExportRequest
from app.services.analysis import analyze
from app.services.persistence import latest_batch

INK = "1D1D1B"
PETROL = "224957"
TEAL = "379B8C"
YELLOW = "E5AD29"
ORANGE = "EA782E"
RED = "E43023"
IVORY = "F7F5EF"


def create_export(db: Session, request: ExportRequest, actor: User) -> BytesIO:
    analysis = analyze(db, request.query_date, request.price_list_code)
    batch = latest_batch(db)
    workbook = Workbook()
    results = workbook.active
    results.title = "Resultados"
    observations = workbook.create_sheet("Observaciones")
    metadata = workbook.create_sheet("Metadatos")

    result_headers = [
        "Fecha consulta", "Lista", "Producto", "Descripción", "Sucursal",
        "Costo vigente", "Vigencia costo", "Precio original", "Vigencia precio",
        "Margen ideal %", "Conductor", "Precio simulado", "Ganancia $", "Ganancia %",
        "Diferencia $", "Diferencia p.p.", "Termómetro", "Estado", "Advertencias",
    ]
    results.append(result_headers)
    for row in analysis.rows:
        simulation_input = request.simulations.get(row.product_code)
        calculated = None
        if simulation_input is not None and not row.simulation_blocked:
            calculated = simulate(
                cost=simulation_input.cost,
                driver=Driver(simulation_input.driver),
                driver_value=simulation_input.driver_value,
                ideal_percent=simulation_input.ideal_percent,
                source_inactive=simulation_input.source_inactive,
                source_unknown=simulation_input.source_unknown,
            )
        results.append([
            request.query_date,
            row.price_list_name,
            row.product_code,
            row.description,
            row.branch_code,
            row.cost.value,
            row.cost.valid_from,
            row.price.value,
            row.price.valid_from,
            row.ideal_percent,
            simulation_input.driver if simulation_input else None,
            decimal_to_str(calculated.price) if calculated else None,
            decimal_to_str(calculated.gain_amount) if calculated else row.actual_gain_amount,
            decimal_to_str(calculated.gain_percent) if calculated else row.actual_gain_percent,
            decimal_to_str(calculated.gap_amount) if calculated else None,
            decimal_to_str(calculated.gap_percentage_points) if calculated else None,
            calculated.thermometer.value if calculated else "neutral",
            row.data_status,
            ", ".join(row.warnings),
        ])

    observations.append(["Tipo", "Severidad", "Hoja", "Clave", "Explicación", "Filas", "Valores"])
    if batch:
        for issue in db.scalars(select(QualityIssue).where(QualityIssue.batch_id == batch.id)).all():
            observations.append([
                issue.issue_type, issue.severity, issue.sheet_name, issue.business_key,
                issue.explanation, ", ".join(map(str, issue.source_rows)), ", ".join(str(v) for v in issue.values),
            ])

    metadata.append(["Campo", "Valor"])
    metadata.append(["Lote", batch.id if batch else ""])
    metadata.append(["Archivo", batch.filename if batch else ""])
    metadata.append(["Fecha de consulta", request.query_date])
    metadata.append(["Lista", analysis.price_list.description])
    metadata.append(["Exportado por", actor.email])
    metadata.append(["Fecha de exportación", datetime.now(UTC).isoformat()])
    metadata.append(["Uso", "Archivo analítico; no apto para carga automática en TOTVS."])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=PETROL)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="center")
        for column in sheet.columns:
            letter = column[0].column_letter
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
            sheet.column_dimensions[letter].width = max(width, 12)

    db.add(AuditEvent(actor_id=actor.id, action="export.created", entity_type="import_batch", entity_id=batch.id if batch else None, details={"date": str(request.query_date), "price_list": request.price_list_code, "rows": len(analysis.rows)}))
    db.commit()
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer

