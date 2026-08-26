"""Pruebas del importador (brief §9 · Datos)."""

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services.importer import parse_workbook

VIG = date(2026, 7, 20)


def _sheet(wb, title, headers, rows, *, first=False):
    ws = wb.active if first else wb.create_sheet(title)
    if first:
        ws.title = title
    ws.append(headers)
    for row in rows:
        ws.append(row)
    return ws


@pytest.fixture
def workbook_path(tmp_path: Path) -> Path:
    wb = Workbook()
    _sheet(
        wb,
        "Mapeo_Listas",
        ["Cod. Tabla", "Descripcion"],
        [[1, "Franquicias Cordoba"]],
        first=True,
    )
    _sheet(
        wb,
        "BD_LP",
        ["Sucursal", "Cod. Tabla", "Cod.Producto", "Precio Venta", "Vigencia", "Activo"],
        [
            [101, 1, "P1", 1000, VIG, "Si"],          # correcto
            [101, 1, "P2", 0, VIG, "Si"],             # precio cero
            [101, 1, "P3", None, VIG, ""],            # precio vacío + estado desconocido
            [101, 1, "P4", 500, VIG, "Si"],           # duplicado idéntico (par)
            [101, 1, "P4", 500, VIG, "Si"],
            [101, 1, "P5", 700, VIG, "Si"],           # duplicado conflictivo (par)
            [101, 1, "P5", 900, VIG, "Si"],
            [101, 1, "P6", 1200, VIG, "No"],          # inactivo
        ],
    )
    _sheet(
        wb,
        "SB1",
        ["Sucursal", "Codigo", "Descripcion", "Costo Estand", "Vigencia", "Desactivado?"],
        [
            [1, "P1", "Producto uno", 800, VIG, "No"],
            [1, "P5", "Producto cinco", 600, VIG, "No"],
        ],
    )
    _sheet(
        wb,
        "Margen_teorico",
        ["Lista", "Código", "Descripción", "Margen"],
        [
            ["Franquicias Cordoba", "P1", "Producto uno", "25,0%"],
            ["Franquicias Cordoba", "varios", "Varios", "30,0%"],
        ],
    )
    path = tmp_path / "base.xlsx"
    wb.save(path)
    return path


def _types(parsed):
    return {issue.issue_type for issue in parsed.issues}


def test_parse_counts_rows(workbook_path):
    parsed = parse_workbook(workbook_path)
    assert parsed.summary["price_rows"] == 8
    assert parsed.summary["cost_rows"] == 2
    assert parsed.summary["margin_rows"] == 2
    assert parsed.summary["price_lists"] == 1


def test_detects_zero_and_missing_price(workbook_path):
    types = _types(parse_workbook(workbook_path))
    assert "zero_price" in types
    assert "missing_price" in types


def test_detects_duplicates(workbook_path):
    parsed = parse_workbook(workbook_path)
    types = _types(parsed)
    assert "identical_duplicate" in types
    assert "conflicting_duplicate" in types
    assert parsed.summary["conflicts"] >= 1


def test_detects_status_flags(workbook_path):
    types = _types(parse_workbook(workbook_path))
    assert "inactive_source" in types
    assert "unknown_source_status" in types


def test_ambiguous_objective_flagged_not_applied(workbook_path):
    parsed = parse_workbook(workbook_path)
    assert "objective_mapping_ambiguous" in _types(parsed)
    ambiguous = [m for m in parsed.margins if m["is_ambiguous"]]
    assert len(ambiguous) == 1
    assert ambiguous[0]["product_code"] == "varios"


def test_missing_sheet_raises(tmp_path):
    wb = Workbook()
    wb.active.title = "Mapeo_Listas"
    wb.active.append(["Cod. Tabla", "Descripcion"])
    path = tmp_path / "incomplete.xlsx"
    wb.save(path)
    with pytest.raises(ValueError, match="Missing required sheets"):
        parse_workbook(path)
