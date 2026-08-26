"""Pruebas de integración (brief §9 · Integración)."""

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.db.base import AuditEvent, PriceFact, Role
from app.db.session import SessionLocal

VIG = date(2026, 7, 20)
QUERY = "2026-08-01"


def _workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapeo_Listas"
    ws.append(["Cod. Tabla", "Descripcion"])
    ws.append([1, "Franquicias Cordoba"])

    lp = wb.create_sheet("BD_LP")
    lp.append(["Sucursal", "Cod. Tabla", "Cod.Producto", "Precio Venta", "Vigencia", "Activo"])
    lp.append([101, 1, "P1", 1000, VIG, "Si"])
    lp.append([101, 1, "P2", 0, VIG, "Si"])

    sb1 = wb.create_sheet("SB1")
    sb1.append(["Sucursal", "Codigo", "Descripcion", "Costo Estand", "Vigencia", "Desactivado?"])
    sb1.append([1, "P1", "Producto uno", 800, VIG, "No"])

    mt = wb.create_sheet("Margen_teorico")
    mt.append(["Lista", "Código", "Descripción", "Margen"])
    mt.append(["Franquicias Cordoba", "P1", "Producto uno", "25,0%"])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _preview(client, header):
    return client.post(
        "/imports/preview",
        files={"file": ("base.xlsx", _workbook_bytes(), "application/octet-stream")},
        headers=header,
    )


# --- Autenticación -----------------------------------------------------------

def test_health_is_open(client):
    assert client.get("/health").json() == {"status": "ok"}


def test_login_allowed(client, make_user):
    make_user("admin@brugali.com.ar", Role.ADMIN_IMPORTER, "clave-de-prueba-larga")
    response = client.post(
        "/auth/login",
        json={"email": "admin@brugali.com.ar", "password": "clave-de-prueba-larga"},
    )
    assert response.status_code == 200
    assert response.json()["user"]["role"] == "admin_importer"


def test_login_denied_wrong_password(client, make_user):
    make_user("admin@brugali.com.ar", Role.ADMIN_IMPORTER, "clave-de-prueba-larga")
    response = client.post(
        "/auth/login",
        json={"email": "admin@brugali.com.ar", "password": "clave-incorrecta"},
    )
    assert response.status_code == 401


def test_analysis_requires_auth(client):
    assert client.get("/analysis", params={"date": QUERY, "price_list": "1"}).status_code == 401


# --- Roles y persistencia ----------------------------------------------------

def test_tester_cannot_preview_or_commit(client, auth_header):
    tester = auth_header("tester@brugali.com.ar", Role.TESTER)
    assert _preview(client, tester).status_code == 403


def test_preview_does_not_persist(client, auth_header):
    admin = auth_header("admin@brugali.com.ar", Role.ADMIN_IMPORTER)
    response = _preview(client, admin)
    assert response.status_code == 200
    assert response.json()["summary"]["price_rows"] == 2
    # Nada se guardó: sin listas de precios ni PriceFact.
    assert client.get("/price-lists", headers=admin).json() == []
    with SessionLocal() as session:
        assert session.scalar(select(PriceFact)) is None


def test_commit_creates_batch_and_audit(client, auth_header):
    admin = auth_header("admin@brugali.com.ar", Role.ADMIN_IMPORTER)
    preview_id = _preview(client, admin).json()["preview_id"]
    commit = client.post(f"/imports/{preview_id}/commit", headers=admin)
    assert commit.status_code == 200
    assert commit.json()["summary"]["price_rows"] == 2
    lists = client.get("/price-lists", headers=admin).json()
    assert any(item["code"] == "1" for item in lists)
    with SessionLocal() as session:
        events = session.scalars(
            select(AuditEvent).where(AuditEvent.action == "import.committed")
        ).all()
        assert len(events) == 1
        # La auditoría no guarda la planilla completa: sólo metadatos acotados.
        assert set(events[0].details) == {"filename", "sha256", "summary"}


def test_tester_cannot_commit_admin_preview(client, auth_header):
    admin = auth_header("admin@brugali.com.ar", Role.ADMIN_IMPORTER)
    tester = auth_header("tester@brugali.com.ar", Role.TESTER)
    preview_id = _preview(client, admin).json()["preview_id"]
    assert client.post(f"/imports/{preview_id}/commit", headers=tester).status_code == 403


# --- Análisis, simulación, exportación --------------------------------------

@pytest.fixture
def committed(client, auth_header):
    admin = auth_header("admin@brugali.com.ar", Role.ADMIN_IMPORTER)
    preview_id = _preview(client, admin).json()["preview_id"]
    client.post(f"/imports/{preview_id}/commit", headers=admin)
    return admin


def test_analysis_returns_rows_and_counts(client, committed):
    response = client.get("/analysis", params={"date": QUERY, "price_list": "1"}, headers=committed)
    assert response.status_code == 200
    body = response.json()
    codes = {row["product_code"] for row in body["rows"]}
    assert {"P1", "P2"} <= codes
    p1 = next(row for row in body["rows"] if row["product_code"] == "P1")
    assert Decimal(p1["cost"]["value"]) == Decimal("800")
    assert Decimal(p1["price"]["value"]) == Decimal("1000")
    assert Decimal(p1["ideal_percent"]) == Decimal("25.0")


def test_saved_simulation_does_not_touch_sources(client, committed):
    payload = {
        "product_code": "P1",
        "price_list_code": "1",
        "query_date": QUERY,
        "cost": "800",
        "ideal_percent": "25",
        "driver": "gain_percent",
        "driver_value": "30",
        "source_inactive": False,
        "source_unknown": False,
    }
    response = client.post("/simulations/save", json=payload, headers=committed)
    assert response.status_code == 200
    assert response.json()["thermometer"] == "green"
    # El precio vigente original no cambió.
    analysis = client.get(
        "/analysis", params={"date": QUERY, "price_list": "1"}, headers=committed
    ).json()
    p1 = next(row for row in analysis["rows"] if row["product_code"] == "P1")
    assert Decimal(p1["price"]["value"]) == Decimal("1000")


def test_negative_price_simulation_rejected(client, committed):
    payload = {
        "product_code": "P1",
        "price_list_code": "1",
        "query_date": QUERY,
        "cost": "800",
        "ideal_percent": None,
        "driver": "price",
        "driver_value": "-5",
        "source_inactive": False,
        "source_unknown": False,
    }
    assert client.post("/simulations/save", json=payload, headers=committed).status_code == 422


def test_export_has_three_sheets(client, committed):
    response = client.post(
        "/exports",
        json={"query_date": QUERY, "price_list_code": "1", "simulations": {}},
        headers=committed,
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Resultados", "Observaciones", "Metadatos"]
    metadata_text = "\n".join(
        str(cell.value) for row in workbook["Metadatos"].iter_rows() for cell in row
    )
    assert "no apto para carga automática en TOTVS" in metadata_text
